
from dataclasses import dataclass
from os import PathLike
from typing import Any, Dict, List, Optional, Sequence
from openpyxl.cell import Cell
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter
import re


@dataclass
class ReportField:
    name: str
    title: str
    width: int
    align: str
    number_format: Optional[str] = None


class StringField(ReportField):
    ALIGNMENT = "left"

    def __init__(self, name, title, width):
        super().__init__(
            name,
            title,
            width,
            self.ALIGNMENT)


class NumberField(ReportField):
    ALIGNMENT = "right"
    WIDTH = 16

    def __init__(self, name, title, number_format="0.0000"):
        super().__init__(
            name,
            title,
            self.WIDTH,
            self.ALIGNMENT,
            number_format=number_format)


class ReportWriter:
    """
    XLSX report writer. 
    """
    TYPEFACE = "Calibri"
    FONT_SIZE = 11
    DEFAULT_ALIGNMENT = "left"
    HEADER_ALIGNMENT = "center"

    def __init__(self, filepath: PathLike):
        self.filepath = filepath

        # Create the workbook and remove the default sheet.
        self.workbook = Workbook()
        del self.workbook["Sheet"]

    def write_sheet(
            self,
            name: str,
            fields,
            data: List[Dict[str, Any]]):
        """
        Adds a sheet to the report.

        Params:
            name: The name of the sheet.
            fields: List of field definitions.
            rows: List of rows to include in sheet.
        """
        # Add the sheet to the workbook.
        worksheet = self.workbook.create_sheet(title=name)

        # Add the header row.
        header_row: List[Cell] = list()
        for field in fields:
            cell = Cell(worksheet, value=field.title)
            # pylint: disable=assigning-non-slot
            cell.alignment = Alignment(horizontal=self.HEADER_ALIGNMENT)
            cell.font = Font(
                name=ReportWriter.TYPEFACE,
                size=ReportWriter.FONT_SIZE,
                bold=True)
            header_row.append(cell)
        worksheet.append(header_row)

        # Add the data rows.
        for row_data in data:
            row: List[Cell] = list()
            for field in fields:
                cell = Cell(worksheet, value=row_data[field["name"]])
                # pylint: disable=assigning-non-slot
                cell.alignment = Alignment(
                    horizontal=ReportWriter.DEFAULT_ALIGNMENT)
                cell.alignment = Alignment(horizontal=field.align)
                cell.font = Font(
                    name=ReportWriter.TYPEFACE,
                    size=ReportWriter.FONT_SIZE)
                if field.number_format:
                    cell.number_format = field.number_format
                row.append(cell)
            worksheet.append(row)

        # Apply column widths.
        for i, field in enumerate(fields):
            column_index = get_column_letter(i + 1)
            worksheet.column_dimensions[column_index].width = field["width"]

    def save(self):
        """
        Write the report to an XLSX file.
        """
        self.workbook.save(self.filepath)


class ReportReader:
    """
    XLSX report reader.
    """

    def __init__(self, filepath):
        workbook = load_workbook(filename=filepath, read_only=True)
        self.worksheet = workbook[workbook.sheetnames[0]]

    def load(self):
        """
        Read rows from the first worksheet.

        Returns:
            List of rows from the worksheet.
        """
        fieldnames = self.get_fieldnames()
        rows: List[Dict[str, Any]] = []  # The rows from the worksheet.

        # Iterate over rows until we hit an empty row.
        for cells in self.worksheet.iter_rows(min_row=2):
            # generate dict from the cells
            row = dict()
            for i, f in enumerate(fieldnames):
                row[f] = cells[i].value

            # discard the row and stop iterating if first column is empty.
            if row[fieldnames[0]] is not None:
                rows.append(row)
            else:
                return rows

    def get_fieldnames(self):
        """
        Read fieldnames from the first row.
        """
        fieldnames: List[str] = []
        col_index = 1
        at_end_of_fields = False
        while not at_end_of_fields:
            fieldname = self.worksheet.cell(1, col_index).value
            if fieldname:
                # Replace spaces and hyphens with underscores.
                fieldname = re.sub(r"[/ -]", "_", fieldname)

                # Remove backslashes and periods.
                fieldname = re.sub(r"[\(\)\.]", "", fieldname)

                fieldname = fieldname.lower()
                fieldnames.append(fieldname)
            else:
                at_end_of_fields = True
            col_index += 1
        return fieldnames
